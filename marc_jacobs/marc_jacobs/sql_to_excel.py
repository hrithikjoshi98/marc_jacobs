from openpyxl import Workbook
import pymysql
import datetime

conn = pymysql.connect(
        host='localhost',
        user='root',
        password="actowiz",
        db='casio',
        autocommit=True
    )
cur = conn.cursor()
wb = Workbook()
ws = wb.active

ws['A1'] = 'id'
ws['B1'] = 'store_no'
ws['C1'] = 'name'
ws['D1'] = 'latitude'
ws['E1'] = 'longitude'
ws['F1'] = 'street'
ws['G1'] = 'city'
ws['H1'] = 'state'
ws['I1'] = 'zip_code'
ws['J1'] = 'county'
ws['K1'] = 'phone'
ws['L1'] = 'open_hours'
ws['M1'] = 'url'
ws['N1'] = 'provider'
ws['O1'] = 'category'
ws['P1'] = 'updated_date'
ws['Q1'] = 'country'
ws['R1'] = 'status'
ws['S1'] = 'direction_url'

date = datetime.datetime.now().strftime('%d_%m_%Y')
file_name = 'marcjacobs_' + str(date) + '_usa'

cur.execute(f"""SELECT * FROM {file_name}""")

all_data = cur.fetchall()
for data in all_data:
    # print(data)
    ws.append(data)

wb.save('marcjacobs_' + str(date) + '_usa' + '.xlsx')